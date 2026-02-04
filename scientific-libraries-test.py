#!/usr/bin/env python3
"""Validation checks for the Termux scientific Python environment."""

from __future__ import annotations

import importlib
import sys
import traceback
import argparse


def module_version(module) -> str:
    for attr in ("__version__", "version", "VERSION"):
        if hasattr(module, attr):
            return str(getattr(module, attr))
    return "unknown"


def run_check(name, import_name, fn, failures):
    print(f"[CHECK] {name}")
    try:
        module = importlib.import_module(import_name)
        print(f"  module: {import_name}")
        print(f"  version: {module_version(module)}")
        fn()
        print("  status: OK")
    except Exception as exc:  # pragma: no cover - intended for runtime diagnostics
        failures.append(name)
        print(f"  status: FAIL ({exc})")
        traceback.print_exc(limit=1)


def check_numpy():
    import numpy as np

    a = np.arange(6).reshape(2, 3)
    b = np.arange(6).reshape(3, 2)
    product = a.dot(b)
    assert product.shape == (2, 2)


def check_pandas():
    import pandas as pd

    df = pd.DataFrame({"x": [1, 2, 3], "y": [10, 20, 30]})
    grouped = df.assign(group=["a", "a", "b"]).groupby("group")["y"].sum()
    assert grouped.loc["a"] == 30


def check_matplotlib():
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    x = np.linspace(0, 1, 10)
    y = np.sin(2 * np.pi * x)
    fig = plt.figure()
    ax = fig.add_subplot(111)
    ax.plot(x, y)
    fig.canvas.draw()
    plt.close(fig)


def check_scipy():
    from scipy import linalg
    import numpy as np

    matrix = np.array([[3.0, 2.0], [1.0, 4.0]])
    inv = linalg.inv(matrix)
    identity = matrix @ inv
    assert identity.shape == (2, 2)


def check_sklearn():
    from sklearn.linear_model import LogisticRegression
    import numpy as np

    x = np.array([[0.0], [1.0], [2.0], [3.0]])
    y = np.array([0, 0, 1, 1])
    model = LogisticRegression(random_state=0).fit(x, y)
    pred = model.predict([[1.5]])
    assert pred.shape == (1,)


def check_statsmodels():
    import numpy as np
    import statsmodels.api as sm

    x = np.arange(1, 11)
    y = 1.0 + 2.0 * x
    design = sm.add_constant(x)
    model = sm.OLS(y, design).fit()
    assert len(model.params) == 2


def check_opencv():
    import cv2
    import numpy as np

    image = np.zeros((8, 8, 3), dtype=np.uint8)
    blurred = cv2.GaussianBlur(image, (3, 3), 0)
    assert blurred.shape == image.shape


def check_seaborn():
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import pandas as pd
    import seaborn as sns

    data = pd.DataFrame({"x": [1, 2, 3], "y": [1, 4, 9]})
    fig = plt.figure()
    ax = fig.add_subplot(111)
    sns.lineplot(data=data, x="x", y="y", ax=ax)
    fig.canvas.draw()
    plt.close(fig)


def check_openpyxl():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "ok"
    assert ws["A1"].value == "ok"


def check_jupyterlab():
    import jupyterlab

    assert hasattr(jupyterlab, "__version__")


def parse_args():
    parser = argparse.ArgumentParser(description="Validate installed scientific Python libraries.")
    parser.add_argument(
        "--lite",
        action="store_true",
        help="Validate only the lite profile libraries.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    print(f"Python executable: {sys.executable}")
    print(f"Python version: {sys.version.split()[0]}")

    failures = []
    full_checks = [
        ("NumPy", "numpy", check_numpy),
        ("Pandas", "pandas", check_pandas),
        ("Matplotlib", "matplotlib", check_matplotlib),
        ("SciPy", "scipy", check_scipy),
        ("Scikit-learn", "sklearn", check_sklearn),
        ("Statsmodels", "statsmodels", check_statsmodels),
        ("OpenCV", "cv2", check_opencv),
        ("Seaborn", "seaborn", check_seaborn),
        ("OpenPyXL", "openpyxl", check_openpyxl),
        ("JupyterLab", "jupyterlab", check_jupyterlab),
    ]
    lite_checks = [
        ("NumPy", "numpy", check_numpy),
        ("Pandas", "pandas", check_pandas),
        ("Matplotlib", "matplotlib", check_matplotlib),
        ("JupyterLab", "jupyterlab", check_jupyterlab),
    ]
    checks = lite_checks if args.lite else full_checks

    for name, import_name, fn in checks:
        run_check(name, import_name, fn, failures)

    print("\n[RESULT]")
    if failures:
        print(f"FAILED: {len(failures)} check(s): {', '.join(failures)}")
        return 1

    print("SUCCESS: all checks passed")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
